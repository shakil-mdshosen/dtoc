from flask import Blueprint, session, redirect, url_for, flash, Response
from models import db, Form, Submission, Permission, AuditLog
import json
import csv
from io import StringIO, BytesIO
from forms_api import login_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

export_bp = Blueprint('export', __name__)

@export_bp.route('/form/<int:form_id>/export/csv')
@login_required
def export_csv(form_id):
    form = Form.query.get_or_404(form_id)
    
    perm = Permission.query.filter_by(form_id=form.id, username=session['username']).first()
    if not perm:
        flash("You don't have permission to export this form.", "danger")
        return redirect(url_for('forms.dashboard'))
        
    submissions = Submission.query.filter_by(form_id=form.id).all()
    
    # Log export
    log = AuditLog(action='EXPORT_CSV', details=f"Form {form.id} exported by {session['username']}")
    db.session.add(log)
    db.session.commit()
    
    # Generate CSV using standard library since pandas isn't available locally
    si = StringIO()
    cw = csv.writer(si)
    
    if not submissions:
        cw.writerow(["No submissions found."])
    else:
        # Extract headers from the first submission
        first_data = json.loads(submissions[0].data)
        if 'data' in first_data:
            # Form.io nests data under 'data' key usually
            first_data = first_data['data']
        headers = list(first_data.keys())
        cw.writerow(['ID', 'Submitted At', 'Submitted By'] + headers)
        
        for sub in submissions:
            data = json.loads(sub.data)
            if 'data' in data:
                data = data['data']
            row = [sub.id, sub.submitted_at.isoformat(), sub.submitted_by]
            row.extend([data.get(h, '') for h in headers])
            cw.writerow(row)
            
    output = si.getvalue()
    
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-disposition": f"attachment; filename=form_{form.id}_export.csv"}
    )

@export_bp.route('/form/<int:form_id>/export/json')
@login_required
def export_json(form_id):
    form = Form.query.get_or_404(form_id)
    
    perm = Permission.query.filter_by(form_id=form.id, username=session['username']).first()
    if not perm:
        flash("You don't have permission to export this form.", "danger")
        return redirect(url_for('forms.dashboard'))
        
    submissions = Submission.query.filter_by(form_id=form.id).all()
    
    # Log export
    log = AuditLog(action='EXPORT_JSON', details=f"Form {form.id} exported as JSON by {session['username']}")
    db.session.add(log)
    db.session.commit()
    
    output_data = []
    for sub in submissions:
        data = json.loads(sub.data)
        if 'data' in data:
            data = data['data']
        output_data.append({
            "id": sub.id,
            "submitted_at": sub.submitted_at.isoformat(),
            "submitted_by": sub.submitted_by,
            "data": data
        })
        
    return Response(
        json.dumps(output_data, indent=2),
        mimetype="application/json",
        headers={"Content-disposition": f"attachment; filename=form_{form.id}_export.json"}
    )

@export_bp.route('/form/<int:form_id>/export/excel')
@login_required
def export_excel(form_id):
    form = Form.query.get_or_404(form_id)

    perm = Permission.query.filter_by(form_id=form.id, username=session['username']).first()
    if not perm:
        flash("You don't have permission to export this form.", "danger")
        return redirect(url_for('forms.dashboard'))

    submissions = Submission.query.filter_by(form_id=form.id).all()

    # Log export
    log = AuditLog(action='EXPORT_EXCEL', details=f"Form {form.id} exported as Excel by {session['username']}")
    db.session.add(log)
    db.session.commit()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submissions"

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    if not submissions:
        ws.append(["No submissions found."])
    else:
        first_data = json.loads(submissions[0].data)
        if 'data' in first_data:
            first_data = first_data['data']
        field_headers = list(first_data.keys())
        all_headers = ['ID', 'Submitted At', 'Submitted By'] + field_headers

        for col_idx, header in enumerate(all_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Freeze the header row
        ws.freeze_panes = "A2"

        for sub in submissions:
            data = json.loads(sub.data)
            if 'data' in data:
                data = data['data']
            row = [sub.id, sub.submitted_at.isoformat(), sub.submitted_by]
            row.extend([data.get(h, '') for h in field_headers])
            ws.append(row)

        # Auto-fit column widths (materialise the column into a list to avoid double-iteration)
        for col in ws.columns:
            cells = list(col)
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in cells)
            ws.column_dimensions[cells[0].column_letter].width = min(max_len + 4, 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": f"attachment; filename=form_{form.id}_export.xlsx"}
    )
